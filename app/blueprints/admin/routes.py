from flask import (
    render_template, request, redirect, url_for,
    session, flash, current_app, jsonify,
)
from app.blueprints.admin import admin_bp
from app.extensions import db
from app.models.drink import Drink
from app.models.member import Member
from app.models.transaction import Transaction
from app.models.setting import Setting


def _require_admin():
    if not session.get("admin"):
        return redirect(url_for("admin.login"))


# ------------------------------------------------------------------
# Auth
# ------------------------------------------------------------------

@admin_bp.get("/login")
def login():
    return render_template("admin/login.html")


@admin_bp.post("/login")
def login_post():
    if request.form.get("password") == current_app.config["ADMIN_PASSWORD"]:
        session["admin"] = True
        return redirect(url_for("admin.index"))
    flash("Wrong password.")
    return redirect(url_for("admin.login"))


@admin_bp.get("/logout")
def logout():
    session.pop("admin", None)
    return redirect(url_for("admin.login"))


# ------------------------------------------------------------------
# Dashboard
# ------------------------------------------------------------------

@admin_bp.get("/")
def index():
    if redir := _require_admin():
        return redir
    total_revenue = db.session.query(
        db.func.sum(Transaction.total_cents)
    ).scalar() or 0
    return render_template(
        "admin/index.html",
        total_revenue=total_revenue / 100,
        transaction_count=Transaction.query.count(),
    )


# ------------------------------------------------------------------
# Drinks
# ------------------------------------------------------------------

@admin_bp.get("/drinks")
def drinks():
    if redir := _require_admin():
        return redir
    return render_template("admin/drinks.html", drinks=Drink.query.order_by(Drink.position).all())


@admin_bp.post("/drinks")
def drinks_create():
    if redir := _require_admin():
        return redir
    drink = Drink(
        name=request.form["name"],
        price_cents=int(float(request.form["price"]) * 100),
        position=Drink.query.count(),
    )
    db.session.add(drink)
    db.session.commit()
    return redirect(url_for("admin.drinks"))


@admin_bp.post("/drinks/<int:drink_id>/edit")
def drinks_edit(drink_id):
    if redir := _require_admin():
        return redir
    drink = Drink.query.get_or_404(drink_id)
    drink.name = request.form["name"]
    drink.price_cents = int(float(request.form["price"]) * 100)
    drink.available = "available" in request.form
    drink.position = int(request.form.get("position", drink.position))
    db.session.commit()
    return redirect(url_for("admin.drinks"))


@admin_bp.post("/drinks/reorder")
def drinks_reorder():
    if redir := _require_admin():
        return redir
    order = request.get_json(silent=True) or []  # list of drink ids in new order
    for position, drink_id in enumerate(order):
        drink = Drink.query.get(drink_id)
        if drink:
            drink.position = position
    db.session.commit()
    return "", 204


@admin_bp.post("/drinks/<int:drink_id>/delete")
def drinks_delete(drink_id):
    if redir := _require_admin():
        return redir
    db.session.delete(Drink.query.get_or_404(drink_id))
    db.session.commit()
    return redirect(url_for("admin.drinks"))


# ------------------------------------------------------------------
# Members / RFID import
# ------------------------------------------------------------------

def _parse_rfid_excel(file_obj):
    """Parse uploaded Excel file, return list of rfid_hex strings."""
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    for name in wb.sheetnames:
        if name.lower() == "chips":
            ws = wb[name]
            break

    rows = list(ws.iter_rows(values_only=True))
    rfid_col = None
    header_row = None

    for ri, row in enumerate(rows[:5]):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        for alias in ("rfid hid", "rfid_id", "rfid", "uid", "karten-id", "card_id"):
            if alias in row_lower:
                rfid_col = row_lower.index(alias)
                header_row = ri
                break
        if rfid_col is not None:
            break

    if rfid_col is None:
        raise ValueError("RFID-Spalte nicht gefunden; erwartet: 'RFID HID', 'rfid_id', 'rfid' o.ä.")

    results = []
    for row in rows[header_row + 1:]:
        raw = str(row[rfid_col]).strip() if rfid_col < len(row) and row[rfid_col] is not None else ""
        if not raw or raw.lower() == "none":
            continue
        try:
            rfid_hex = format(int(raw), "08X")
        except ValueError:
            rfid_hex = raw.upper()
        results.append(rfid_hex)
    return results


@admin_bp.get("/members")
def members():
    if redir := _require_admin():
        return redir
    all_members = Member.query.order_by(Member.name).all()
    return render_template("admin/members.html", members=all_members)


@admin_bp.post("/members/import")
def members_import():
    if redir := _require_admin():
        return redir
    f = request.files.get("excel_file")
    if not f or not f.filename:
        flash("Keine Datei ausgewählt.")
        return redirect(url_for("admin.members"))
    try:
        rows = _parse_rfid_excel(f.stream)
    except Exception as exc:
        flash(f"Fehler beim Lesen der Datei: {exc}")
        return redirect(url_for("admin.members"))

    added = skipped = 0
    for rfid_hex in rows:
        if Member.query.filter_by(rfid_uid=rfid_hex).first():
            skipped += 1
            continue
        db.session.add(Member(name=rfid_hex, rfid_uid=rfid_hex))
        added += 1
    db.session.commit()
    flash(f"{added} importiert, {skipped} bereits vorhanden.")
    return redirect(url_for("admin.members"))


@admin_bp.post("/members/<int:member_id>/topup")
def members_topup(member_id):
    if redir := _require_admin():
        return redir
    member = Member.query.get_or_404(member_id)
    try:
        amount_cents = int(float(request.form["amount"]) * 100)
    except (KeyError, ValueError):
        flash("Ungültiger Betrag.")
        return redirect(url_for("admin.members"))
    member.balance_cents += amount_cents
    db.session.commit()
    flash(f"Guthaben von {member.name} aufgeladen.")
    return redirect(url_for("admin.members"))


@admin_bp.post("/members/<int:member_id>/delete")
def members_delete(member_id):
    if redir := _require_admin():
        return redir
    member = Member.query.get_or_404(member_id)
    member.active = False
    db.session.commit()
    flash(f"{member.name} deaktiviert.")
    return redirect(url_for("admin.members"))


# ------------------------------------------------------------------
# Analytics
# ------------------------------------------------------------------

def _uid_stats():
    """Shared helper: returns (all_txs, uid_rows) for analytics routes."""
    from collections import defaultdict
    all_txs = Transaction.query.order_by(Transaction.created_at).all()
    uid_stats: dict[str, dict] = defaultdict(lambda: {"count": 0, "total_cents": 0})
    for tx in all_txs:
        uid = tx.rfid_uid or ""
        uid_stats[uid]["count"] += 1
        uid_stats[uid]["total_cents"] += tx.total_cents
    uid_rows = sorted(
        [{"uid": uid, **stats} for uid, stats in uid_stats.items()],
        key=lambda r: r["total_cents"],
        reverse=True,
    )
    return all_txs, uid_rows


@admin_bp.get("/analytics")
def analytics():
    if redir := _require_admin():
        return redir
    transactions = Transaction.query.order_by(Transaction.created_at.desc()).limit(100).all()
    _, uid_rows = _uid_stats()
    return render_template("admin/analytics.html", transactions=transactions, uid_rows=uid_rows)


@admin_bp.get("/charts")
def charts():
    if redir := _require_admin():
        return redir
    from collections import defaultdict
    all_txs, uid_rows = _uid_stats()

    # Revenue per day (last 30 days with data)
    day_revenue: dict[str, int] = defaultdict(int)
    seen: list[str] = []
    for tx in all_txs:
        d = tx.created_at.strftime("%d.%m.")
        day_revenue[d] += tx.total_cents
        if d not in seen:
            seen.append(d)
    chart_days        = seen[-30:]
    chart_day_revenue = [round(day_revenue[d] / 100, 2) for d in chart_days]

    # Transactions per hour of day
    hour_count = [0] * 24
    for tx in all_txs:
        hour_count[tx.created_at.hour] += 1
    chart_hours = [f"{h:02d}:00" for h in range(24)]

    # Top 10 UIDs
    top_uids          = uid_rows[:10]
    chart_uid_labels  = [r["uid"] for r in top_uids]
    chart_uid_revenue = [round(r["total_cents"] / 100, 2) for r in top_uids]
    chart_uid_count   = [r["count"] for r in top_uids]

    return render_template(
        "admin/charts.html",
        chart_days=chart_days,
        chart_day_revenue=chart_day_revenue,
        chart_hours=chart_hours,
        hour_count=hour_count,
        chart_uid_labels=chart_uid_labels,
        chart_uid_revenue=chart_uid_revenue,
        chart_uid_count=chart_uid_count,
    )


# ------------------------------------------------------------------
# Sync (quick trigger from dashboard)
# ------------------------------------------------------------------

@admin_bp.post("/sync")
def sync_now():
    if redir := _require_admin():
        return redir
    from app import sync_service
    sync_service.run_now()
    flash("Sync ausgelöst.")
    return redirect(url_for("admin.index"))


# ------------------------------------------------------------------
# Settings
# ------------------------------------------------------------------

_SETTING_KEYS = ["CLOUD_URL", "CLOUD_API_KEY", "BAR_UID"]


@admin_bp.get("/settings")
def settings():
    if redir := _require_admin():
        return redir
    from app import sync_service
    values = {k: Setting.get(k, current_app.config.get(k, "")) for k in _SETTING_KEYS}
    unsynced = Transaction.query.filter_by(synced=False).count()
    status = sync_service.status() if sync_service else {}
    return render_template(
        "admin/settings.html",
        values=values,
        unsynced=unsynced,
        sync_status=status,
    )


@admin_bp.post("/settings")
def settings_save():
    if redir := _require_admin():
        return redir
    for key in _SETTING_KEYS:
        val = request.form.get(key, "").strip()
        Setting.set(key, val)
    flash("Einstellungen gespeichert.")
    return redirect(url_for("admin.settings"))


@admin_bp.post("/settings/sync")
def settings_sync():
    if redir := _require_admin():
        return redir
    from app import sync_service
    result = sync_service.run_now_sync()
    return jsonify(result)


@admin_bp.post("/settings/test")
def settings_test():
    if redir := _require_admin():
        return redir
    url     = Setting.get("CLOUD_URL",     current_app.config.get("CLOUD_URL", ""))
    api_key = Setting.get("CLOUD_API_KEY", current_app.config.get("CLOUD_API_KEY", ""))
    if not url or not api_key:
        return jsonify({"ok": False, "msg": "URL oder API-Key nicht konfiguriert."})
    try:
        import requests as _requests
        resp = _requests.get(f"{url.rstrip('/')}/health", timeout=5)
        if resp.status_code == 200:
            return jsonify({"ok": True, "msg": f"Verbunden ({resp.json().get('utc', '')})"})
        return jsonify({"ok": False, "msg": f"HTTP {resp.status_code}"})
    except Exception as exc:
        return jsonify({"ok": False, "msg": str(exc)})
